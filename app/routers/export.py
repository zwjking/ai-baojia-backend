"""
报价导出接口 - GET /api/quote/{quote_id}/export?format=pdf|xlsx

V7 (陈浩 2026-07-13):
  - 君哥要求: 详细项目报价 + 导出 PDF / Excel
  - 数据源: quotes 表 (quote_json 字段存的是 breakdown_v4 + items)
  - 格式: reportlab 4.x 生成 PDF (中文需嵌入字体) / openpyxl 3.x 生成 xlsx

字体方案:
  - PDF 标题/正文用 Helvetica (英文) + STSong-Light (中文 CID, 系统自带)
  - 备选: 注册本地中文字体 (msyh.ttc)
"""
from __future__ import annotations

import io
import json
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, status
from fastapi.responses import Response, StreamingResponse

from app.models.database import Quote, SessionLocal

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/quote", tags=["export"])


# ============== 取数据 ==============
def _load_quote(quote_id: int) -> dict:
    """从 DB 取报价(包含完整 quote_json 解析后的 breakdown_v4 + items)"""
    db = SessionLocal()
    try:
        q = db.query(Quote).filter(Quote.id == quote_id).first()
        if not q:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error": "not_found", "hint": f"报价 ID {quote_id} 不存在"},
            )
        # 解析 quote_json
        try:
            quote_data = json.loads(q.quote_json)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail={"error": "db_corrupt", "hint": f"quote_json 解析失败: {e}"},
            )
        # 解析 survey_json
        try:
            survey_data = json.loads(q.survey_json)
        except Exception:
            survey_data = {}
        return {
            "id": q.id,
            "user_id": q.user_id,
            "source": q.source,
            "generated_at": q.generated_at.isoformat() if q.generated_at else None,
            "total": q.total_amount,
            "ml_correction": quote_data.get("ml_correction"),
            "items_count": quote_data.get("items_count"),
            "breakdown_summary": quote_data.get("breakdown", {}),
            "survey": survey_data,
            # quote_json 里没存 items 详情,得重算 - 但 fallback V4 不会重算
            # 解决方案: 在 quote.py 写库时把完整 items + breakdown_v4 存到 quote_json
        }
    finally:
        db.close()


# ============== 触发一次新报价拿完整数据 ==============
def _recompute_full(quote_id: int) -> dict:
    """用存量的 survey_json 重新调一次 fallback 拿完整 breakdown_v4 + items"""
    from app.models.schemas import QuoteRequest, GradeEnum, PackEnum, DistrictEnum, BrandTierEnum
    from app.services.fallback import _load_prices, compute_fallback

    info = _load_quote(quote_id)
    s = info["survey"]
    # 重构 QuoteRequest
    def _to_brand_tier(v):
        if not v:
            return None
        try:
            return BrandTierEnum(v)
        except Exception:
            return None
    try:
        req = QuoteRequest(
            area=float(s["area"]),
            layout=s["layout"],
            grade=GradeEnum(s.get("grade", "中档")),
            pack=PackEnum(s.get("pack", "全包")),
            style=s.get("style", "现代"),
            special=s.get("special", []) or [],
            district=DistrictEnum(s.get("district", "蜀山区")),
            contact="13800138000",  # 占位
            rooms=s.get("rooms"),
            floor=s.get("floor"),
            has_elevator=s.get("has_elevator"),
            demolition_wall_area=s.get("demolition_wall_area"),
            demolition_build_area=s.get("demolition_build_area"),
            brand_tier_tile=_to_brand_tier(s.get("brand_tier_tile")),
            brand_tier_floor=_to_brand_tier(s.get("brand_tier_floor")),
            brand_tier_cabinet=_to_brand_tier(s.get("brand_tier_cabinet")),
            brand_tier_bathroom=_to_brand_tier(s.get("brand_tier_bathroom")),
        )
    except Exception as e:
        logger.warning("重构 QuoteRequest 失败: %s, 用基础信息导出", e)
        return {**info, "breakdown_v4": None, "items": []}

    prices = _load_prices()
    try:
        resp, _ = compute_fallback(req)
        return {
            **info,
            "breakdown_v4": resp.breakdown_v4.model_dump() if resp.breakdown_v4 else None,
            "items": [it.model_dump() for it in resp.items],
            "total": resp.total,
            "demolition_cost": resp.demolition_cost,
            "material_brand_tier": resp.material_brand_tier,
        }
    except Exception as e:
        logger.warning("重算失败: %s, 用基础信息导出", e)
        return {**info, "breakdown_v4": None, "items": []}


# ============== PDF 导出 ==============
@router.get("/{quote_id}/export")
async def export_quote(
    quote_id: int,
    format: str = Query("pdf", regex="^(pdf|xlsx)$", description="导出格式: pdf 或 xlsx"),
):
    """
    V7 导出接口: PDF / Excel
    - PDF: reportlab 生成, 含项目信息 + 5 类分类 + 详细 SKU 明细表
    - xlsx: openpyxl 生成, 含封面 + 5 分类 sheet + 全部 items
    """
    data = _recompute_full(quote_id)
    if format == "pdf":
        content = _render_pdf(data)
        filename = f"wenjun_quote_{quote_id}.pdf"
        media_type = "application/pdf"
    else:
        content = _render_xlsx(data)
        filename = f"wenjun_quote_{quote_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


def _render_pdf(data: dict) -> bytes:
    """用 reportlab 生成 PDF 报价单"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    # 注册中文字体 (reportlab 自带 CID 字体, 无需本地文件)
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        cn_font = "STSong-Light"
    except Exception:
        cn_font = "Helvetica"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="闻君 AI 装修报价单",
    )
    styles = getSampleStyleSheet()

    # 自定义样式
    title_style = ParagraphStyle(
        "cn_title", parent=styles["Title"],
        fontName=cn_font, fontSize=24, leading=30,
        textColor=colors.HexColor("#1f2937"),
    )
    h1_style = ParagraphStyle(
        "cn_h1", parent=styles["Heading1"],
        fontName=cn_font, fontSize=16, leading=22,
        textColor=colors.HexColor("#ea580c"),
    )
    h2_style = ParagraphStyle(
        "cn_h2", parent=styles["Heading2"],
        fontName=cn_font, fontSize=13, leading=18,
        textColor=colors.HexColor("#111827"),
    )
    body_style = ParagraphStyle(
        "cn_body", parent=styles["BodyText"],
        fontName=cn_font, fontSize=10, leading=14,
    )
    cell_style = ParagraphStyle(
        "cn_cell", parent=body_style, fontSize=9, leading=12,
    )

    story = []

    # 标题
    story.append(Paragraph("闻君 AI 装修报价单", title_style))
    story.append(Paragraph(
        f"报价 ID: {data['id']} | 来源: {data.get('source', '-')} | "
        f"生成时间: {data.get('generated_at', '-')}",
        body_style,
    ))
    story.append(Spacer(1, 0.6 * cm))

    # 项目信息
    s = data.get("survey", {})
    info_rows = [
        ["建筑面积", f"{s.get('area', '-')} ㎡", "户型", s.get("layout", "-")],
        ["装修档次", s.get("grade", "-"), "包工模式", s.get("pack", "-")],
        ["装修风格", s.get("style", "-"), "所在区域", s.get("district", "-")],
        ["房间", s.get("rooms") or "-", "楼层", str(s.get("floor") or "-")],
        ["电梯", "有" if s.get("has_elevator") else ("无" if s.get("has_elevator") is False else "-"),
         "拆墙/砌墙", f"{s.get('demolition_wall_area', 0)} ㎡ / {s.get('demolition_build_area', 0)} ㎡"],
        ["地砖/地板/橱柜/卫浴品牌档次",
         f"{s.get('brand_tier_tile', '-')} / {s.get('brand_tier_floor', '-')} / "
         f"{s.get('brand_tier_cabinet', '-')} / {s.get('brand_tier_bathroom', '-')}",
         "特殊需求", ", ".join(s.get("special", []) or []) or "-"],
    ]
    story.append(Paragraph("一、项目信息", h1_style))
    info_table = Table(
        [[Paragraph(str(c), cell_style) for c in row] for row in info_rows],
        colWidths=[3.5 * cm, 5 * cm, 3 * cm, 5 * cm],
    )
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f9fafb")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.6 * cm))

    # 总价
    total = data.get("total", 0)
    total_table = Table(
        [[Paragraph(f"AI 核算总价(含税): <b>¥ {total:,.2f}</b>", ParagraphStyle(
            "total", parent=body_style, fontSize=18, leading=24,
            textColor=colors.HexColor("#dc2626"), alignment=1,
        ))]],
        colWidths=[16.5 * cm],
    )
    total_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7ed")),
        ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#ea580c")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(total_table)
    story.append(Spacer(1, 0.6 * cm))

    # 5 类分类
    bv4 = data.get("breakdown_v4") or {}
    cat_map = {
        "main_material": "主材",
        "auxiliary": "辅材",
        "labor": "人工",
        "management": "管理费",
        "tax": "税金",
    }
    story.append(Paragraph("二、5 项费用分类", h1_style))
    cat_rows = [["分类", "金额(元)", "占比"]]
    total_sum = total or 1
    for k, name in cat_map.items():
        cat = bv4.get(k) or {}
        amt = cat.get("total", 0)
        pct = (amt / total_sum * 100) if total_sum else 0
        cat_rows.append([name, f"¥ {amt:,.2f}", f"{pct:.1f}%"])
    cat_table = Table(
        [[Paragraph(str(c), cell_style) for c in row] for row in cat_rows],
        colWidths=[5 * cm, 6 * cm, 5.5 * cm],
    )
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), cn_font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(cat_table)
    story.append(Spacer(1, 0.6 * cm))

    # 详细 SKU 明细
    story.append(Paragraph("三、详细项目报价(SKU 级别)", h1_style))
    items = data.get("items") or []
    if items:
        item_rows = [["序号", "项目名称", "类别", "规格/品牌", "单位", "工程量", "单价(元)", "合价(元)"]]
        for i, it in enumerate(items, 1):
            item_rows.append([
                str(i),
                it.get("name", "-"),
                it.get("category", "-"),
                f"{(it.get('brand') or '-')} / {(it.get('spec') or '-')}",
                it.get("unit", "-"),
                f"{float(it.get('quantity', 0)):.2f}",
                f"{float(it.get('unit_price', 0)):,.2f}",
                f"{float(it.get('total', 0)):,.2f}",
            ])
        item_table = Table(
            [[Paragraph(str(c), cell_style) for c in row] for row in item_rows],
            colWidths=[1 * cm, 4.5 * cm, 1.8 * cm, 3.5 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm],
        )
        item_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), cn_font),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9ca3af")),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(item_table)
    else:
        story.append(Paragraph("(无明细数据)", body_style))

    # 备注
    story.append(Spacer(1, 0.6 * cm))
    note_style = ParagraphStyle(
        "note", parent=body_style, fontSize=9, textColor=colors.HexColor("#6b7280"),
    )
    story.append(Paragraph(
        "<b>价格说明:</b> 本报价基于合肥本地 2026 年 7 月建材价格库 + 顾工 V4 算账公式,"
        f"实际价格会因主材品牌、施工面积、拆改量浮动 ±5%。<b>半包</b>不含家具、家电、软装。",
        note_style,
    ))
    story.append(Paragraph(
        "<b>免责声明:</b> 本报价为 AI 参考价,不作决策依据。准确报价请联系闻君工长上门量房。",
        note_style,
    ))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"闻君聊装修 · 合肥本地 11 年 · ICP 备 12345678 号 · "
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ParagraphStyle("foot", parent=note_style, alignment=1, fontSize=8),
    ))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def _render_xlsx(data: dict) -> bytes:
    """用 openpyxl 生成 Excel 报价单"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 默认 sheet -> 项目概览
    ws_overview = wb.active
    ws_overview.title = "项目概览"

    # 样式
    title_font = Font(name="微软雅黑", size=18, bold=True, color="1F2937")
    h1_font = Font(name="微软雅黑", size=14, bold=True, color="EA580C")
    h2_font = Font(name="微软雅黑", size=11, bold=True, color="111827")
    body_font = Font(name="微软雅黑", size=10)
    big_total_font = Font(name="微软雅黑", size=20, bold=True, color="DC2626")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # 标题
    ws_overview.merge_cells("A1:F1")
    ws_overview["A1"] = "闻君 AI 装修报价单"
    ws_overview["A1"].font = title_font
    ws_overview["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_overview.row_dimensions[1].height = 36

    # 副标题
    ws_overview.merge_cells("A2:F2")
    ws_overview["A2"] = f"报价 ID: {data['id']} | 来源: {data.get('source', '-')} | 生成时间: {data.get('generated_at', '-')}"
    ws_overview["A2"].font = body_font
    ws_overview["A2"].alignment = Alignment(horizontal="center")

    # 项目信息
    s = data.get("survey", {})
    row = 4
    ws_overview[f"A{row}"] = "一、项目信息"
    ws_overview[f"A{row}"].font = h1_font
    row += 1
    info_pairs = [
        ("建筑面积", f"{s.get('area', '-')} ㎡"),
        ("户型", s.get("layout", "-")),
        ("装修档次", s.get("grade", "-")),
        ("包工模式", s.get("pack", "-")),
        ("装修风格", s.get("style", "-")),
        ("所在区域", s.get("district", "-")),
        ("房间", s.get("rooms") or "-"),
        ("楼层", str(s.get("floor") or "-")),
        ("电梯", "有" if s.get("has_elevator") else ("无" if s.get("has_elevator") is False else "-")),
        ("拆墙", f"{s.get('demolition_wall_area', 0)} ㎡"),
        ("砌墙", f"{s.get('demolition_build_area', 0)} ㎡"),
        ("地砖品牌档次", s.get("brand_tier_tile", "-")),
        ("地板品牌档次", s.get("brand_tier_floor", "-")),
        ("橱柜品牌档次", s.get("brand_tier_cabinet", "-")),
        ("卫浴品牌档次", s.get("brand_tier_bathroom", "-")),
        ("特殊需求", ", ".join(s.get("special", []) or []) or "-"),
    ]
    for k, v in info_pairs:
        ws_overview[f"A{row}"] = k
        ws_overview[f"A{row}"].font = h2_font
        ws_overview[f"A{row}"].fill = alt_fill
        ws_overview[f"A{row}"].border = border
        ws_overview.merge_cells(f"B{row}:F{row}")
        ws_overview[f"B{row}"] = v
        ws_overview[f"B{row}"].font = body_font
        ws_overview[f"B{row}"].border = border
        row += 1

    row += 1
    # 总价
    ws_overview.merge_cells(f"A{row}:F{row}")
    ws_overview[f"A{row}"] = f"AI 核算总价(含税): ¥ {data.get('total', 0):,.2f}"
    ws_overview[f"A{row}"].font = big_total_font
    ws_overview[f"A{row}"].fill = orange_fill
    ws_overview[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws_overview[f"A{row}"].border = border
    ws_overview.row_dimensions[row].height = 48
    row += 2

    # 5 类分类
    ws_overview[f"A{row}"] = "二、5 项费用分类"
    ws_overview[f"A{row}"].font = h1_font
    row += 1
    bv4 = data.get("breakdown_v4") or {}
    cat_map = [
        ("main_material", "主材"),
        ("auxiliary", "辅材"),
        ("labor", "人工"),
        ("management", "管理费"),
        ("tax", "税金"),
    ]
    total = data.get("total", 0) or 1
    headers = ["分类", "金额(元)", "占比"]
    for i, h in enumerate(headers, 1):
        c = ws_overview.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1
    for k, name in cat_map:
        cat = bv4.get(k) or {}
        amt = cat.get("total", 0)
        pct = (amt / total * 100) if total else 0
        ws_overview.cell(row=row, column=1, value=name).font = body_font
        ws_overview.cell(row=row, column=2, value=amt).font = body_font
        ws_overview.cell(row=row, column=2).number_format = "¥ #,##0.00"
        ws_overview.cell(row=row, column=3, value=pct / 100).font = body_font
        ws_overview.cell(row=row, column=3).number_format = "0.0%"
        for i in range(1, 4):
            ws_overview.cell(row=row, column=i).border = border
            if i > 1:
                ws_overview.cell(row=row, column=i).alignment = Alignment(horizontal="right")
        row += 1

    # 设置列宽
    for i, w in enumerate([14, 14, 14, 14, 14, 14], 1):
        ws_overview.column_dimensions[get_column_letter(i)].width = w

    # 明细 sheet: 所有 items
    ws_items = wb.create_sheet("详细项目报价")
    headers = ["序号", "项目名称", "类别", "品牌", "规格", "单位", "工程量", "单价(元)", "合价(元)"]
    for i, h in enumerate(headers, 1):
        c = ws_items.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    items = data.get("items") or []
    for idx, it in enumerate(items, 1):
        r = idx + 1
        ws_items.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws_items.cell(row=r, column=2, value=it.get("name", "-"))
        ws_items.cell(row=r, column=3, value=it.get("category", "-"))
        ws_items.cell(row=r, column=4, value=it.get("brand") or "-")
        ws_items.cell(row=r, column=5, value=it.get("spec") or "-")
        ws_items.cell(row=r, column=6, value=it.get("unit", "-")).alignment = Alignment(horizontal="center")
        ws_items.cell(row=r, column=7, value=float(it.get("quantity", 0))).number_format = "0.00"
        ws_items.cell(row=r, column=8, value=float(it.get("unit_price", 0))).number_format = "¥ #,##0.00"
        ws_items.cell(row=r, column=9, value=float(it.get("total", 0))).number_format = "¥ #,##0.00"
        for i in range(1, 10):
            cell = ws_items.cell(row=r, column=i)
            cell.font = body_font
            cell.border = border
            if idx % 2 == 0:
                cell.fill = alt_fill
    # 合计行
    sum_row = len(items) + 2
    ws_items.cell(row=sum_row, column=8, value="合计").font = h2_font
    ws_items.cell(row=sum_row, column=8).alignment = Alignment(horizontal="right")
    total_amt = sum(float(it.get("total", 0)) for it in items)
    ws_items.cell(row=sum_row, column=9, value=total_amt).font = h2_font
    ws_items.cell(row=sum_row, column=9).number_format = "¥ #,##0.00"
    ws_items.cell(row=sum_row, column=9).fill = orange_fill
    ws_items.cell(row=sum_row, column=9).border = border

    # 列宽
    widths = [6, 30, 10, 16, 24, 8, 10, 14, 16]
    for i, w in enumerate(widths, 1):
        ws_items.column_dimensions[get_column_letter(i)].width = w
    ws_items.freeze_panes = "A2"

    # 5 类分类各建一个 sheet
    for k, name in cat_map:
        cat = bv4.get(k) or {}
        cat_items = cat.get("items") or []
        if not cat_items:
            continue
        ws_cat = wb.create_sheet(name)
        for i, h in enumerate(headers, 1):
            c = ws_cat.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center")
        for idx, it in enumerate(cat_items, 1):
            r = idx + 1
            ws_cat.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws_cat.cell(row=r, column=2, value=it.get("name", "-"))
            ws_cat.cell(row=r, column=3, value=it.get("category", name))
            ws_cat.cell(row=r, column=4, value=it.get("brand") or "-")
            ws_cat.cell(row=r, column=5, value=it.get("spec") or "-")
            ws_cat.cell(row=r, column=6, value=it.get("unit", "-")).alignment = Alignment(horizontal="center")
            ws_cat.cell(row=r, column=7, value=float(it.get("quantity", 0))).number_format = "0.00"
            ws_cat.cell(row=r, column=8, value=float(it.get("unit_price", 0))).number_format = "¥ #,##0.00"
            ws_cat.cell(row=r, column=9, value=float(it.get("total", 0))).number_format = "¥ #,##0.00"
            for i in range(1, 10):
                cell = ws_cat.cell(row=r, column=i)
                cell.font = body_font
                cell.border = border
                if idx % 2 == 0:
                    cell.fill = alt_fill
        for i, w in enumerate(widths, 1):
            ws_cat.column_dimensions[get_column_letter(i)].width = w
        ws_cat.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    buf.close()
    return xlsx_bytes
